# Django
from django.core.files.base import ContentFile

# Python
import openpyxl
import pytz
from PIL import Image
from io import BytesIO


def copy_to_excel(inbox_messages):
    # Creating new document
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Writing column headers
    sheet['A1'] = 'Sender'
    sheet['B1'] = 'Recipients'
    sheet['C1'] = 'Subject'
    sheet['D1'] = 'Message'
    sheet['E1'] = 'Time'

    # Writing data from inbox_messages in Excel
    for index, message in enumerate(inbox_messages):
        row = index + 2
        sheet.cell(row=row, column=1).value = str(message.sender.email)
        recipients = ', '.join(str(recipient)
                               for recipient in message.recipients.all())
        sheet.cell(row=row, column=2).value = recipients
        sheet.cell(row=row, column=3).value = message.subject
        sheet.cell(row=row, column=4).value = message.body
        tz_aware_dt = message.timestamp
        tz = pytz.timezone('Asia/Almaty')
        tz_aware_dt = tz_aware_dt.astimezone(tz)
        tz_naive_dt = tz_aware_dt.replace(tzinfo=None)

        sheet.cell(row=row, column=5).value = tz_naive_dt

    # Saving file to Excel
    workbook.save('inbox_messages.xlsx')


def copy_outbox_to_excel(outbox_messages):
    # Creating new document
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Writing column headers
    sheet['A1'] = 'Sender'
    sheet['B1'] = 'Recipients'
    sheet['C1'] = 'Subject'
    sheet['D1'] = 'Message'
    sheet['E1'] = 'Time'

    # Writing data from outbox_messages in Excel
    for index, message in enumerate(outbox_messages):
        # Starting from the second line (first contains headers)
        row = index + 2
        sheet.cell(row=row, column=1).value = str(message.sender.email)
        recipients = ', '.join(str(recipient)
                               for recipient in message.recipients.all())
        sheet.cell(row=row, column=2).value = recipients
        sheet.cell(row=row, column=3).value = message.subject
        sheet.cell(row=row, column=4).value = message.body
        tz_aware_dt = message.timestamp
        tz = pytz.timezone('Asia/Almaty')
        tz_aware_dt = tz_aware_dt.astimezone(tz)
        tz_naive_dt = tz_aware_dt.replace(tzinfo=None)

        sheet.cell(row=row, column=5).value = tz_naive_dt

    # Saving file to Excel
    workbook.save('outbox_messages.xlsx')


def copy_outbox_external_to_excel(outbox_messages):
    # Creating new document
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Writing column headers
    sheet['A1'] = 'To mail'
    sheet['B1'] = 'Additional mail'
    sheet['C1'] = 'From'
    sheet['D1'] = 'Subject'
    sheet['E1'] = 'Message'
    sheet['F1'] = 'Time'

    # Writing data from outbox_external_messages in Excel
    for index, message in enumerate(outbox_messages):
        # Starting from the second line (first contains headers)
        row = index + 2
        sheet.cell(row=row, column=1).value = message.recipient
        sheet.cell(row=row, column=2).value = message.additional_recipient
        sheet.cell(row=row, column=3).value = message.sender.email
        sheet.cell(row=row, column=4).value = message.subject
        sheet.cell(row=row, column=5).value = message.message
        tz_aware_dt = message.timestamp
        tz = pytz.timezone('Asia/Almaty')
        tz_aware_dt = tz_aware_dt.astimezone(tz)
        tz_naive_dt = tz_aware_dt.replace(tzinfo=None)

        sheet.cell(row=row, column=6).value = tz_naive_dt

    # Saving file to Excel
    workbook.save('outbox_external_messages.xlsx')


def process_and_save_photo(photo):
    # Photo processing and saving
    image = Image.open(photo)
    image = image.convert('RGB')
    image.thumbnail((150, 150))
    buffer = BytesIO()
    image.save(buffer, format='JPEG')
    photo_file = buffer.getvalue()

    return ContentFile(photo_file)


# encrypting messsages in Ceasars's method
def encrypt_caesar(plaintext, shift):
    encrypted_text = ""
    for char in plaintext:
        if char.isalpha():
            start = ord('a') if char.islower() else ord('A')
            encrypted_char = chr((ord(char) - start + shift) % 26 + start)
            encrypted_text += encrypted_char
        else:
            encrypted_text += char
    return encrypted_text


def decrypt_caesar(ciphertext, shift):
    return encrypt_caesar(ciphertext, -shift)
